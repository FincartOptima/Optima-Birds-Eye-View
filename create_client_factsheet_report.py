from __future__ import annotations

import csv
import io
import math
import re
import sys

# Bump from Python's 1000 default. Higher (50k+) overflows the OS thread
# stack on Windows and crashes the worker without a catchable error.
sys.setrecursionlimit(5000)
from bisect import bisect_right
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen.canvas import Canvas as PdfCanvas
from reportlab.lib.colors import HexColor

from custodian_statement import CustodianStatement, CorpusDeposit

REPORT_DATE = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)


def set_report_date(date: datetime) -> None:
    """Align the report/valuation date to the snapshot CSV's holding date, so
    XIRR and the displayed report date match the date the NAVs belong to."""
    global REPORT_DATE
    REPORT_DATE = date.replace(hour=0, minute=0, second=0, microsecond=0)


def get_report_date() -> datetime:
    return REPORT_DATE


CATEGORY_BY_ISIN = {
    "INF109KC1RH9": "Indian Equity",
    "INF200K01RJ1": "Indian Equity",
    "INF109K01X40": "Indian Equity",
    "INF247L01CW5": "Indian Equity",
    "INF740K01QA7": "Indian Equity",
    "INF00XX01747": "Indian Equity",
    "INF194KB1AL4": "Indian Equity",
    "INF205K013T3": "Indian Equity",
    "INF846K01X06": "Foreign Equity",
    "INF846K01Y39": "Foreign Equity",
    "INF090I01JR0": "Foreign Equity",
    "INF090I01IZ5": "Foreign Equity",
    "INF174K01MP6": "Gold",
    "INF109K01U92": "Gold",
    "INF846K01DI3": "Debt",
    "INF277K017Q3": "Cash Fund",
    "INF174K01NE8": "Cash Fund",
    "INF754K01LB7": "Foreign Equity",
}

CATEGORY_ORDER = ["Indian Equity", "Foreign Equity", "Gold", "Debt", "Cash Fund", "Only Cash"]


@dataclass
class Transaction:
    source_sheet: str
    source_row: int
    statement_date: datetime
    transaction_type: str
    ucc: str
    client_name: str
    scheme_name: str
    folio_no: str
    isin: str
    amount: float
    units: float | None
    value_date: datetime | str | None
    allocation_date: datetime | str | None


@dataclass
class FailedTransaction:
    """A trade-master row whose Value column (fixed position, column K) was
    blank. Excluded from every holdings/XIRR/dashboard calculation."""
    source_sheet: str
    source_row: int
    ucc: str
    client_name: str
    scheme_name: str
    amount_units: float | None
    value_date: datetime | str | None
    reason: str


@dataclass
class Holding:
    isin: str
    scheme_name: str
    category: str
    units: float = 0.0
    cost_value: float = 0.0
    realized_pl: float = 0.0
    current_nav: float | None = None
    current_value: float = 0.0
    unrealized_pl: float = 0.0
    total_pl: float = 0.0
    nav_source: str = ""


@dataclass
class ClientReport:
    client_name: str
    ucc: str
    initial_investment: float
    initial_date: datetime | None
    transactions: list[Transaction] = field(default_factory=list)
    holdings: list[Holding] = field(default_factory=list)
    only_cash: float = 0.0
    cost_value: float = 0.0
    current_value: float = 0.0
    unrealized_pl: float = 0.0
    realized_pl: float = 0.0
    total_pl: float = 0.0
    xirr: float | None = None
    simple_return: float | None = None
    contribution: float = 0.0
    net_gain: float = 0.0
    benchmark_current_value: float | None = None
    benchmark_xirr: float | None = None
    category_rows: list[dict[str, Any]] = field(default_factory=list)
    top_holdings: list[Holding] = field(default_factory=list)
    performance_rows: list[tuple[datetime, float | None, float | None]] = field(default_factory=list)
    custodian: CustodianStatement | None = None


# ---------------------------------------------------------------------------
# Text / data helpers
# ---------------------------------------------------------------------------

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_header(value: Any) -> str:
    return clean_text(value).lower()


def to_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            return None
    return None


def as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                pass
    return None


def parse_sheet_date(sheet_name: str) -> datetime | None:
    # An optional leading "R" marks a dedicated redemption sheet (e.g.
    # "R08072026"). The date itself is still the trailing 8 digits.
    match = re.fullmatch(r"[Rr]?(\d{8})", sheet_name)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%d%m%Y")
    except ValueError:
        return None


def is_redemption_sheet_name(sheet_name: str) -> bool:
    return bool(re.fullmatch(r"[Rr]\d{8}", sheet_name))


def header_index(headers: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for pos, header in enumerate(headers):
        key = clean_header(header)
        if key and key not in index:
            index[key] = pos
    return index


def row_value(row: list[Any], headers: list[Any], *aliases: str) -> Any:
    index = header_index(headers)
    for alias in aliases:
        pos = index.get(clean_header(alias))
        if pos is not None and pos < len(row):
            return row[pos]
    return None


def infer_category(isin: str, scheme_name: str) -> str:
    if isin in CATEGORY_BY_ISIN:
        return CATEGORY_BY_ISIN[isin]
    text = scheme_name.lower()
    if any(word in text for word in ["u.s.", " us ", "asian", "china", "global", "technology equity"]):
        return "Foreign Equity"
    if "gold" in text:
        return "Gold"
    if any(word in text for word in ["bond", "debt", "gilt"]):
        return "Debt"
    if any(word in text for word in ["liquid", "arbitrage", "money market", "overnight"]):
        return "Cash Fund"
    return "Indian Equity"


# ---------------------------------------------------------------------------
# Data readers
# ---------------------------------------------------------------------------

def validate_trade_master(workbook) -> list[str]:
    """Return a list of format errors for the trade master workbook.
    An empty list means the file is valid."""
    errors: list[str] = []
    if "Master" not in workbook.sheetnames:
        errors.append("Missing required 'Master' sheet. The Trade Master file must contain a sheet named 'Master'.")
    else:
        sheet = workbook["Master"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        idx = header_index(headers)
        if idx.get("ucc code") is None:
            errors.append("The 'Master' sheet is missing the 'UCC Code' column header.")
        if idx.get("client name") is None:
            errors.append("The 'Master' sheet is missing the 'Client Name' column header.")

    date_sheets = [s for s in workbook.sheetnames if parse_sheet_date(s) is not None]
    if not date_sheets:
        errors.append(
            "No transaction sheets found. The file must contain at least one sheet "
            "named as DDMMYYYY (e.g. '08072026') or RDDMMYYYY for redemptions."
        )
    return errors


def validate_client_csv(file_path: Path) -> list[str]:
    """Return a list of format errors for the client holdings CSV.
    An empty list means the file is valid."""
    errors: list[str] = []
    required = {"client code", "isin", "quantity", "market value"}
    try:
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                errors.append("The CSV file appears to be empty or has no header row.")
                return errors
            found = {h.strip().lower() for h in reader.fieldnames if h}
            missing = required - found
            if missing:
                nice = ", ".join(sorted(m.title() for m in missing))
                errors.append(f"The client CSV is missing required column(s): {nice}.")
    except UnicodeDecodeError:
        errors.append("The client file is not a valid CSV — it may be an Excel file renamed to .csv.")
    except Exception as exc:
        errors.append(f"Could not read the client CSV: {exc}")
    return errors


def read_master(workbook) -> dict[str, dict[str, Any]]:
    if "Master" not in workbook.sheetnames:
        return {}
    sheet = workbook["Master"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = header_index(headers)
    master: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_values = list(row)
        ucc_pos = index.get("ucc code")
        client_pos = index.get("client name")
        initial_pos = index.get("initial investment")
        if ucc_pos is None or client_pos is None:
            continue
        ucc = clean_text(row_values[ucc_pos] if ucc_pos < len(row_values) else None)
        client_name = clean_text(row_values[client_pos] if client_pos < len(row_values) else None)
        if not ucc or not client_name:
            continue
        master[ucc] = {
            "client_name": client_name,
            "initial_investment": to_number(row_values[initial_pos] if initial_pos is not None else None) or 0.0,
            "initial_date": as_datetime(row_values[index.get("date of opening")] if index.get("date of opening") is not None else None)
            or as_datetime(row_values[index.get("form received date")] if index.get("form received date") is not None else None),
        }
    return master


# Fixed 0-based column positions for the failed-transaction Reason text
# (columns C, Z, V). These stay positional -- the "settlement value" check
# itself is resolved by header name (see _find_column below) so it keeps
# working across sheets whose layout/column order differs.
_REASON_COL_IDX_C = 2   # column C
_REASON_COL_IDX_Z = 25  # column Z
_REASON_COL_IDX_V = 21  # column V


def _cell(row: list[Any], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def _find_column(headers: list[Any], *aliases: str) -> int | None:
    index = header_index(headers)
    for alias in aliases:
        pos = index.get(clean_header(alias))
        if pos is not None:
            return pos
    return None


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def read_transactions(
    workbook, master: dict[str, dict[str, Any]]
) -> tuple[list[Transaction], list[FailedTransaction]]:
    transactions: list[Transaction] = []
    failed_transactions: list[FailedTransaction] = []
    for sheet in workbook.worksheets:
        statement_date = parse_sheet_date(sheet.title)
        if statement_date is None:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = list(rows[0])
        transaction_type = "Redemption" if is_redemption_sheet_name(sheet.title) else "Subscription"
        has_type_column = _find_column(headers, "Type of Instruction") is not None
        for row_number, row_tuple in enumerate(rows[1:], start=2):
            row = list(row_tuple)
            first_cell = clean_header(row[0] if row else None)
            if "redemption" in first_cell and first_cell != "redemption":
                transaction_type = "Redemption"
                continue
            if first_cell == "ucc":
                headers = row
                has_type_column = _find_column(headers, "Type of Instruction") is not None
                if any(clean_header(header) == "redemption" for header in headers):
                    transaction_type = "Redemption"
                continue
            ucc = clean_text(row_value(row, headers, "UCC"))
            client_name = clean_text(row_value(row, headers, "Client Name"))
            scheme_name = clean_text(row_value(row, headers, "Scheme Name"))
            isin = clean_text(row_value(row, headers, "ISIN Code"))
            if not (re.fullmatch(r"CRFM\d+", ucc, re.I) and client_name and scheme_name):
                continue
            if ucc in master:
                client_name = master[ucc]["client_name"]

            # A per-row "Type of Instruction" column (when present) is the
            # most reliable signal -- it only ever pushes toward Redemption,
            # never resets an already-active redemption section back to
            # Subscription, so it can't disturb the older marker-row logic.
            row_type = transaction_type
            if has_type_column:
                instruction = clean_header(row_value(row, headers, "Type of Instruction"))
                if instruction == "redemption":
                    row_type = "Redemption"

            # Resolve the money amount and the settlement-pending value.
            # Older sheets carry both in separate columns: "Amount/Units"
            # (money, always filled up front) and a units-ish column that's
            # blank until the order settles. Newer sheets (e.g. the
            # "R"-prefixed redemption format) have no combined Amount/Units
            # column at all -- units-to-redeem is filled up front instead,
            # and the "Value" column (₹ realized) is what's blank until
            # settlement, so that becomes both the amount and the check.
            amount_col = _find_column(headers, "Amount/Units")
            if amount_col is not None:
                amount = to_number(_cell(row, amount_col))
                settle_col = _find_column(headers, "UNITS", "Units", "Value")
            else:
                settle_col = _find_column(headers, "Value")
                amount = to_number(_cell(row, settle_col)) if settle_col is not None else None

            settle_cell = _cell(row, settle_col) if settle_col is not None else None
            is_failed = settle_col is None or _is_blank(settle_cell)

            if is_failed:
                reason_parts = [
                    clean_text(_cell(row, _REASON_COL_IDX_C)),
                    clean_text(_cell(row, _REASON_COL_IDX_Z)),
                    clean_text(_cell(row, _REASON_COL_IDX_V)),
                ]
                failed_transactions.append(
                    FailedTransaction(
                        source_sheet=sheet.title,
                        source_row=row_number,
                        ucc=ucc,
                        client_name=client_name,
                        scheme_name=scheme_name,
                        amount_units=amount,
                        value_date=row_value(row, headers, "Value Date"),
                        reason=" ".join(part for part in reason_parts if part),
                    )
                )
                continue

            if amount is None:
                continue

            transactions.append(
                Transaction(
                    source_sheet=sheet.title,
                    source_row=row_number,
                    statement_date=statement_date,
                    transaction_type=row_type,
                    ucc=ucc,
                    client_name=client_name,
                    scheme_name=scheme_name,
                    folio_no=clean_text(row_value(row, headers, "Folio No")),
                    isin=isin,
                    amount=amount,
                    units=to_number(row_value(row, headers, "UNITS", "Units", "Value")),
                    value_date=row_value(row, headers, "Value Date"),
                    allocation_date=row_value(row, headers, "Child Allocation", "Child Allocation ", "Redemption"),
                )
            )
    return transactions, failed_transactions


def read_current_navs(file_path: Path) -> tuple[dict[str, float], dict[str, str]]:
    if not file_path.exists():
        return {}, {}
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = header_index(headers)
    navs: dict[str, float] = {}
    categories: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_values = list(row)
        isin_pos = index.get("isin")
        nav_pos = index.get("current nav")
        category_pos = index.get("category")
        if isin_pos is None or nav_pos is None:
            continue
        isin = clean_text(row_values[isin_pos] if isin_pos < len(row_values) else None)
        nav = to_number(row_values[nav_pos] if nav_pos < len(row_values) else None)
        if isin and nav is not None:
            navs[isin] = nav
            if category_pos is not None and category_pos < len(row_values):
                category = clean_text(row_values[category_pos])
                if category:
                    categories[isin] = category
    return navs, categories


def read_client_file_csv(file_path: Path) -> tuple[dict[str, float], dict[str, str]]:
    """Read a custodian holdings-snapshot CSV (e.g. 'client file.csv').

    The snapshot has one row per (client, holding) with the live NAV in the
    'Unit Price' column. We extract a {ISIN: current_nav} map so the snapshot
    can act as the source of current valuations (replacing Current_NAVs.xlsx).
    Categories are left to CATEGORY_BY_ISIN / infer_category, so we return an
    empty category-override dict.
    """
    navs: dict[str, float] = {}
    if not file_path.exists():
        return navs, {}
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            isin = clean_text(row.get("ISIN"))
            nav = to_number(row.get("Unit Price"))
            if isin and nav is not None and nav > 0:
                navs[isin] = nav
    return navs, {}


def read_snapshot_details(
    file_path: Path,
) -> tuple[datetime | None, dict[str, dict[str, dict[str, Any]]], dict[str, float]]:
    """Per-client detail from the custodian holdings-snapshot CSV.

    Returns (holding_date, {ucc: {isin: {'units', 'cost', 'scheme'}}}, {ucc: cash}).
    The snapshot's quantities and costs are the custodian's official allotted
    figures as of the holding date; they override trade-master-derived units
    so every tab of the webpage values the same portfolio on the same date.
    Cash counts only the literal uninvested CASH rows.
    """
    holdings: dict[str, dict[str, dict[str, Any]]] = {}
    cash: dict[str, float] = {}
    holding_date: datetime | None = None
    if not file_path.exists():
        return None, holdings, cash
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            ucc = clean_text(row.get("Client code"))
            if not ucc:
                continue
            if holding_date is None:
                try:
                    holding_date = datetime.strptime(clean_text(row.get("Holding Date")), "%d/%m/%Y")
                except ValueError:
                    pass
            symbol = clean_text(row.get("SYMBOLCODE")).upper()
            scheme = clean_text(row.get("SYMBOLNAME"))
            isin = clean_text(row.get("ISIN"))
            mkt = to_number(row.get("Market Value")) or 0.0
            if symbol == "CASH" or (not isin and scheme.lower() == "cash"):
                cash[ucc] = cash.get(ucc, 0.0) + mkt
                continue
            units = to_number(row.get("QUANTITY"))
            if not isin or units is None:
                continue
            holdings.setdefault(ucc, {})[isin] = {
                "units": units,
                "cost": to_number(row.get("Total Cost")) or 0.0,
                "scheme": scheme,
            }
    return holding_date, holdings, cash


def latest_observed_navs(transactions: list[Transaction]) -> dict[str, tuple[float, datetime]]:
    navs: dict[str, tuple[float, datetime]] = {}
    for transaction in transactions:
        if not transaction.isin or not transaction.units:
            continue
        nav = transaction.amount / transaction.units
        current = navs.get(transaction.isin)
        if current is None or transaction.statement_date >= current[1]:
            navs[transaction.isin] = (nav, transaction.statement_date)
    return navs


# ---------------------------------------------------------------------------
# XIRR / benchmark helpers
# ---------------------------------------------------------------------------

def xnpv(rate: float, cashflows: list[tuple[datetime, float]]) -> float:
    start = min(date for date, _ in cashflows)
    total = 0.0
    for date, amount in cashflows:
        years = (date - start).days / 365.0
        total += amount / ((1.0 + rate) ** years)
    return total


def xirr(cashflows: list[tuple[datetime, float]]) -> float | None:
    valid_cashflows = [(date, amount) for date, amount in cashflows if abs(amount) > 0.000001]
    if not valid_cashflows:
        return None
    if not any(amount > 0 for _, amount in valid_cashflows) or not any(amount < 0 for _, amount in valid_cashflows):
        return None
    low = -0.9999
    high = 10.0
    low_value = xnpv(low, valid_cashflows)
    high_value = xnpv(high, valid_cashflows)
    expansion_count = 0
    while low_value * high_value > 0 and expansion_count < 20:
        high *= 2
        high_value = xnpv(high, valid_cashflows)
        expansion_count += 1
    if low_value * high_value > 0:
        return None
    for _ in range(100):
        mid = (low + high) / 2
        mid_value = xnpv(mid, valid_cashflows)
        if abs(mid_value) < 0.000001:
            return mid
        if low_value * mid_value <= 0:
            high = mid
            high_value = mid_value
        else:
            low = mid
            low_value = mid_value
    return (low + high) / 2


def price_on_or_before(prices: list[tuple[datetime, float]], date: datetime) -> float | None:
    dates = [item[0] for item in prices]
    pos = bisect_right(dates, date) - 1
    if pos < 0:
        return None
    return prices[pos][1]


def benchmark_value_and_xirr(
    transactions: list[Transaction],
    prices: list[tuple[datetime, float]],
    report_date: datetime,
    initial_investment: float = 0.0,
    initial_date: datetime | None = None,
) -> tuple[float | None, float | None]:
    if not prices:
        return None, None
    if initial_investment > 0:
        start_date = initial_date or (min(item.statement_date for item in transactions) if transactions else report_date)
        start_price = price_on_or_before(prices, start_date)
        final_price = price_on_or_before(prices, report_date) or prices[-1][1]
        if not start_price:
            return None, None
        current_value = initial_investment * final_price / start_price
        return current_value, xirr([(start_date, -initial_investment), (report_date, current_value)])
    if not transactions:
        return None, None
    units = 0.0
    cashflows: list[tuple[datetime, float]] = []
    for transaction in sorted(transactions, key=lambda item: (item.statement_date, item.source_sheet, item.source_row)):
        price = price_on_or_before(prices, transaction.statement_date)
        if not price:
            continue
        if transaction.transaction_type == "Redemption":
            units -= transaction.amount / price
            cashflows.append((transaction.statement_date, transaction.amount))
        else:
            units += transaction.amount / price
            cashflows.append((transaction.statement_date, -transaction.amount))
    final_price = price_on_or_before(prices, report_date) or prices[-1][1]
    current_value = units * final_price
    if current_value > 0:
        cashflows.append((report_date, current_value))
    return current_value, xirr(cashflows)




def make_benchmark_series(
    transactions: list[Transaction],
    prices: list[tuple[datetime, float]],
    report_date: datetime,
) -> list[tuple[datetime, float]]:
    if not transactions:
        return []
    sorted_transactions = sorted(transactions, key=lambda item: (item.statement_date, item.source_row))
    start_date = min(item.statement_date for item in sorted_transactions)
    units = 0.0
    tx_index = 0
    series: list[tuple[datetime, float]] = []
    for date, price in prices:
        if date < start_date:
            continue
        if date > report_date:
            break
        while tx_index < len(sorted_transactions) and sorted_transactions[tx_index].statement_date <= date:
            transaction = sorted_transactions[tx_index]
            tx_price = price_on_or_before(prices, transaction.statement_date)
            if tx_price:
                if transaction.transaction_type == "Redemption":
                    units -= transaction.amount / tx_price
                else:
                    units += transaction.amount / tx_price
            tx_index += 1
        series.append((date, units * price))
    return series


def make_client_series(
    transactions: list[Transaction],
    holdings: dict[str, Holding],
    report_date: datetime,
) -> list[tuple[datetime, float]]:
    if not transactions:
        return []
    nav_history: dict[str, list[tuple[datetime, float]]] = defaultdict(list)
    for transaction in transactions:
        if transaction.units:
            nav_history[transaction.isin].append((transaction.statement_date, transaction.amount / transaction.units))
    for isin in nav_history:
        nav_history[isin].sort(key=lambda item: item[0])
        final_nav = holdings.get(isin).current_nav if holdings.get(isin) else None
        if final_nav:
            nav_history[isin].append((report_date, final_nav))
    event_dates = sorted({transaction.statement_date for transaction in transactions} | {report_date})
    holding_units: dict[str, float] = defaultdict(float)
    series: list[tuple[datetime, float]] = []
    sorted_transactions = sorted(transactions, key=lambda item: (item.statement_date, item.source_row))
    tx_index = 0
    for date in event_dates:
        while tx_index < len(sorted_transactions) and sorted_transactions[tx_index].statement_date <= date:
            transaction = sorted_transactions[tx_index]
            if transaction.units:
                if transaction.transaction_type == "Redemption":
                    holding_units[transaction.isin] -= transaction.units
                else:
                    holding_units[transaction.isin] += transaction.units
            tx_index += 1
        value = 0.0
        for isin, units in holding_units.items():
            history = nav_history.get(isin, [])
            dates = [item[0] for item in history]
            pos = bisect_right(dates, date) - 1
            if pos >= 0:
                value += units * history[pos][1]
        series.append((date, value))
    return series


# ---------------------------------------------------------------------------
# Build client reports
# ---------------------------------------------------------------------------

def build_client_reports(
    master: dict[str, dict[str, Any]],
    transactions: list[Transaction],
    bse_prices: list[tuple[datetime, float]],
    current_navs: dict[str, float],
    category_overrides: dict[str, str],
    custodian_data: dict[str, CustodianStatement] | None = None,
    snapshot_holdings: dict[str, dict[str, dict[str, Any]]] | None = None,
    snapshot_cash: dict[str, float] | None = None,
) -> list[ClientReport]:
    custodian_data = custodian_data or {}
    snapshot_holdings = snapshot_holdings or {}
    snapshot_cash = snapshot_cash or {}
    transactions_by_ucc: dict[str, list[Transaction]] = defaultdict(list)
    for transaction in transactions:
        transactions_by_ucc[transaction.ucc].append(transaction)
    latest_navs = latest_observed_navs(transactions)
    reports: list[ClientReport] = []
    for ucc, master_row in sorted(master.items(), key=lambda item: item[1]["client_name"].lower()):
        client_transactions = sorted(
            transactions_by_ucc.get(ucc, []),
            key=lambda item: (item.statement_date, item.source_sheet, item.source_row),
        )
        custodian = custodian_data.get(ucc)
        report = ClientReport(
            client_name=master_row["client_name"],
            ucc=ucc,
            initial_investment=master_row["initial_investment"],
            initial_date=master_row.get("initial_date"),
            transactions=client_transactions,
            custodian=custodian,
        )
        holdings_by_isin: dict[str, Holding] = {}
        for transaction in client_transactions:
            if transaction.isin not in holdings_by_isin:
                category = category_overrides.get(transaction.isin) or infer_category(transaction.isin, transaction.scheme_name)
                holdings_by_isin[transaction.isin] = Holding(
                    isin=transaction.isin,
                    scheme_name=transaction.scheme_name,
                    category=category,
                )
            holding = holdings_by_isin[transaction.isin]
            units = transaction.units or 0.0
            if not units:
                fallback_nav = current_navs.get(transaction.isin)
                if fallback_nav is None and transaction.isin in latest_navs:
                    fallback_nav = latest_navs[transaction.isin][0]
                if fallback_nav:
                    units = transaction.amount / fallback_nav
                else:
                    units = transaction.amount
            if transaction.transaction_type == "Redemption":
                avg_cost = holding.cost_value / holding.units if holding.units else 0.0
                cost_removed = min(units, holding.units) * avg_cost if units else min(transaction.amount, holding.cost_value)
                holding.units -= units
                holding.cost_value -= cost_removed
                holding.realized_pl += transaction.amount - cost_removed
            else:
                holding.units += units
                holding.cost_value += transaction.amount
        # The custodian snapshot's official units/cost (as of the holding
        # date) override the trade-master-derived figures, so the factsheet
        # values exactly the portfolio the custodian reports.
        for isin, snap_row in snapshot_holdings.get(ucc, {}).items():
            if isin not in holdings_by_isin:
                category = category_overrides.get(isin) or infer_category(isin, snap_row["scheme"])
                holdings_by_isin[isin] = Holding(
                    isin=isin, scheme_name=snap_row["scheme"], category=category,
                )
            holding = holdings_by_isin[isin]
            holding.units = snap_row["units"]
            holding.cost_value = snap_row["cost"]
        for holding in holdings_by_isin.values():
            if holding.isin in current_navs:
                holding.current_nav = current_navs[holding.isin]
                holding.nav_source = "Current_NAVs.xlsx"
            elif holding.isin in latest_navs:
                holding.current_nav = latest_navs[holding.isin][0]
                holding.nav_source = "Latest transaction NAV"
            elif holding.units and holding.cost_value:
                holding.current_nav = holding.cost_value / holding.units
                holding.nav_source = "Cost fallback"
            holding.current_value = holding.units * holding.current_nav if holding.current_nav else 0.0
            holding.unrealized_pl = holding.current_value - holding.cost_value
            holding.total_pl = holding.unrealized_pl + holding.realized_pl
        report.holdings = sorted(holdings_by_isin.values(), key=lambda item: item.current_value, reverse=True)
        report.realized_pl = sum(holding.realized_pl for holding in report.holdings)
        report.cost_value = sum(holding.cost_value for holding in report.holdings)
        report.current_value = sum(holding.current_value for holding in report.holdings)
        net_fund_investment = sum(
            transaction.amount if transaction.transaction_type == "Subscription" else -transaction.amount
            for transaction in client_transactions
        )
        # Cash precedence: snapshot CSV CASH row (same as-of date as the
        # valuation) > custodian statement cash > master-derived residual.
        if ucc in snapshot_cash:
            report.only_cash = snapshot_cash[ucc]
        elif custodian and custodian.cash is not None:
            report.only_cash = custodian.cash
        else:
            report.only_cash = max(report.initial_investment - net_fund_investment, 0.0)
        report.cost_value += report.only_cash
        report.current_value += report.only_cash
        report.unrealized_pl = report.current_value - report.cost_value
        report.total_pl = report.unrealized_pl + report.realized_pl
        category_values = defaultdict(lambda: {"cost": 0.0, "current": 0.0, "unrealized": 0.0, "realized": 0.0})
        for holding in report.holdings:
            category_values[holding.category]["cost"] += holding.cost_value
            category_values[holding.category]["current"] += holding.current_value
            category_values[holding.category]["unrealized"] += holding.unrealized_pl
            category_values[holding.category]["realized"] += holding.realized_pl
        category_values["Only Cash"]["cost"] += report.only_cash
        category_values["Only Cash"]["current"] += report.only_cash
        report.category_rows = []
        for category in CATEGORY_ORDER:
            row = category_values[category]
            current = row["current"]
            report.category_rows.append(
                {
                    "Category": category,
                    "Cost Value": row["cost"],
                    "Current Value": current,
                    "Allocation %": current / report.current_value if report.current_value else 0.0,
                    "Unrealized P/L": row["unrealized"],
                    "Realized P/L": row["realized"],
                    "Total P/L": row["unrealized"] + row["realized"],
                }
            )
        report.top_holdings = [holding for holding in report.holdings if holding.current_value > 0][:5]

        # Build XIRR cashflows using real deposit dates from the account
        # statement when available; fall back to the Master sheet's single
        # lump-sum figure otherwise.
        effective_initial_investment = report.initial_investment
        effective_initial_date = report.initial_date
        portfolio_cashflows: list[tuple[datetime, float]] = []

        dated_deposits = [d for d in custodian.deposits if d.date <= REPORT_DATE] if custodian else []
        if dated_deposits:
            for dep in dated_deposits:
                portfolio_cashflows.append((dep.date, -dep.amount))
            effective_initial_investment = sum(d.amount for d in dated_deposits)
            effective_initial_date = dated_deposits[0].date
        elif effective_initial_investment > 0:
            start_date = effective_initial_date or (
                min(transaction.statement_date for transaction in client_transactions) if client_transactions else REPORT_DATE
            )
            portfolio_cashflows.append((start_date, -effective_initial_investment))
        else:
            for transaction in client_transactions:
                amount = transaction.amount if transaction.transaction_type == "Redemption" else -transaction.amount
                portfolio_cashflows.append((transaction.statement_date, amount))

        if report.current_value:
            portfolio_cashflows.append((REPORT_DATE, report.current_value))
        report.xirr = xirr(portfolio_cashflows)

        # Invested = total corpus deposited (custodian contribution when
        # available). Gain/Loss and Simple Return both measure against it, so
        # the KPI strip is internally consistent: Gain% == Simple Return.
        report.contribution = effective_initial_investment if effective_initial_investment > 0 else report.cost_value
        report.net_gain = report.current_value - report.contribution
        report.simple_return = (report.net_gain / report.contribution) if report.contribution > 0 else None

        report.benchmark_current_value, report.benchmark_xirr = benchmark_value_and_xirr(
            client_transactions, bse_prices, REPORT_DATE, effective_initial_investment, effective_initial_date
        )
        client_series = make_client_series(client_transactions, holdings_by_isin, REPORT_DATE)
        benchmark_series = make_benchmark_series(client_transactions, bse_prices, REPORT_DATE)
        benchmark_by_date = {date: value for date, value in benchmark_series}
        report.performance_rows = [
            (date, value, benchmark_by_date.get(date)) for date, value in client_series
        ]
        reports.append(report)
    return reports


# =====================================================================
# PDF FACTSHEET GENERATION
# =====================================================================

_PW, _PH = A4  # 595.28, 841.89
_LM = 28
_RM = 28
_CW = _PW - _LM - _RM

_NAVY = HexColor("#14365C")
_DARK_NAVY = HexColor("#0D2440")
_GOLD = HexColor("#C5922E")
_CREAM = HexColor("#F3E9CE")
_WHITE = HexColor("#FFFFFF")
_LIGHT_BG = HexColor("#F4F6F9")
_LIGHT_BLUE = HexColor("#E8EEF6")
_BORDER = HexColor("#C2D0E0")
_TEXT_DARK = HexColor("#2C3E50")
_TEXT_MED = HexColor("#5A6C7E")
_TEXT_LIGHT = HexColor("#8899AA")
_GREEN_VAL = HexColor("#1B7A2F")
_RED_VAL = HexColor("#C0392B")

_PIE_COLORS = ["#1F4E78", "#3A7CA5", "#C5922E", "#5DADE2", "#A0B4C8"]
_BAR_COLOR = "#1F4E78"

_PDF_CAT_DISPLAY = {
    "Indian Equity": "Domestic Equity MF",
    "Foreign Equity": "International Equity",
    "Gold": "Commodities (Gold + Nat Res)",
    "Debt": "Debt Mutual Funds",
    "Cash Fund": "Cash & Equivalent",
    "Only Cash": "Cash & Equivalent",
}


def _fmt_pct(val: float | None, sign: bool = True) -> str:
    if val is None:
        return "N/A"
    pct = val * 100
    if sign and pct >= 0:
        return f"+{pct:.2f}%"
    return f"{pct:.2f}%"


def _pct_color(val: float | None):
    if val is None or val == 0:
        return _TEXT_DARK
    return _GREEN_VAL if val > 0 else _RED_VAL


def _inception(report: ClientReport) -> datetime:
    if report.custodian and report.custodian.deposits:
        return report.custodian.deposits[0].date
    if report.initial_date:
        return report.initial_date
    if report.transactions:
        return min(t.statement_date for t in report.transactions)
    return REPORT_DATE


def _fmt_inr(val: float | None) -> str:
    if val is None or val == 0:
        return "Rs 0"
    if abs(val) >= 1e7:
        return f"Rs {val / 1e7:.2f} Cr"
    if abs(val) >= 1e5:
        return f"Rs {val / 1e5:.2f} L"
    return f"Rs {val:,.0f}"


def generate_client_pdf(report: ClientReport, output_path: Path) -> None:
    """Simple, single-page factsheet — no matplotlib, minimal ReportLab calls."""
    c = PdfCanvas(str(output_path), pagesize=A4)
    _draw_simple_factsheet(c, report)
    c.save()


def _draw_simple_factsheet(c: PdfCanvas, report: ClientReport) -> None:
    inception_dt = _inception(report)
    gl_frac = (report.net_gain / report.contribution) if report.contribution else 0.0

    # ---- Header band -----------------------------------------------------
    y = _PH - 12
    c.setFillColor(_NAVY)
    c.rect(_LM, y - 48, _CW, 48, fill=True, stroke=False)
    c.setFillColor(_GOLD)
    c.rect(_LM, y - 51, _CW, 3, fill=True, stroke=False)
    c.setFillColor(_WHITE)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(_LM + 14, y - 18, "Fincart Direct Mutual Fund Strategic Portfolio")
    c.setFont("Helvetica", 8)
    c.drawString(_LM + 14, y - 32, f"Client: {report.client_name}   |   UCC: {report.ucc}")
    c.drawRightString(_LM + _CW - 14, y - 18, "FACTSHEET")
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(_LM + _CW - 14, y - 32, REPORT_DATE.strftime("%B %Y"))
    y -= 62

    # ---- KPI strip (5 boxes) -------------------------------------------
    box_h = 50
    kpis = [
        ("INVESTED",      _fmt_inr(report.contribution),  "Total corpus deposited", None),
        ("CURRENT VALUE", _fmt_inr(report.current_value), "", None),
        ("GAIN / LOSS",   _fmt_inr(report.net_gain),      _fmt_pct(gl_frac), gl_frac),
        ("SIMPLE RETURN", _fmt_pct(report.simple_return) if report.simple_return is not None else "N/A",
                          "Non-annualised, since inception",
                          report.simple_return),
        ("XIRR (ANNUALISED)", _fmt_pct(report.xirr) if report.xirr is not None else "N/A",
                          f"BSE 500: {_fmt_pct(report.benchmark_xirr) if report.benchmark_xirr is not None else 'N/A'}",
                          report.xirr),
    ]
    box_w = _CW / len(kpis)
    for i, (label, value, sub, color_value) in enumerate(kpis):
        bx = _LM + i * box_w
        c.setFillColor(_LIGHT_BLUE)
        c.rect(bx, y - box_h, box_w, box_h, fill=True, stroke=False)
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.4)
        c.rect(bx, y - box_h, box_w, box_h, fill=False, stroke=True)
        c.setFillColor(_TEXT_MED)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(bx + box_w / 2, y - 12, label)
        c.setFillColor(_pct_color(color_value) if color_value is not None else _NAVY)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(bx + box_w / 2, y - 28, value)
        if sub:
            c.setFillColor(_TEXT_MED)
            c.setFont("Helvetica", 7)
            c.drawCentredString(bx + box_w / 2, y - 42, sub)
    y -= box_h + 12

    # ---- Client / strategy details --------------------------------------
    details = [
        ("Client",           report.client_name),
        ("UCC",              report.ucc),
        ("Inception Date",   inception_dt.strftime("%d %b %Y")),
        ("Report Date",      REPORT_DATE.strftime("%d %b %Y")),
        ("Strategy",         "Multi-Asset MF Portfolio (Direct Plans)"),
        ("Benchmark",        "BSE 500"),
        ("Portfolio Manager","Credent Asset Management  |  SEBI INP000006101"),
    ]
    y = _draw_simple_section(c, y, "PORTFOLIO SUMMARY")
    y = _draw_two_col_table(c, y, details, label_w=140)
    y -= 8

    # ---- Asset allocation table -----------------------------------------
    alloc_rows = [
        (row["Category"],
         _fmt_inr(row["Current Value"]),
         f"{row['Allocation %'] * 100:.2f}%")
        for row in report.category_rows if row["Current Value"] > 0
    ]
    if alloc_rows:
        alloc_rows.append(("TOTAL", _fmt_inr(report.current_value), "100.00%"))
        y = _draw_simple_section(c, y, "ASSET ALLOCATION")
        y = _draw_three_col_table(
            c, y,
            ["Category", "Current Value", "Weight"],
            alloc_rows,
            col_widths=[_CW * 0.5, _CW * 0.3, _CW * 0.2],
            total_last=True,
        )
        y -= 8

    # ---- Top holdings ----------------------------------------------------
    holding_rows = [
        (h.scheme_name[:52],
         _fmt_inr(h.current_value),
         f"{(h.current_value / report.current_value * 100) if report.current_value else 0:.2f}%")
        for h in report.top_holdings
    ]
    if holding_rows:
        y = _draw_simple_section(c, y, "TOP 5 HOLDINGS")
        y = _draw_three_col_table(
            c, y,
            ["Fund", "Current Value", "Weight"],
            holding_rows,
            col_widths=[_CW * 0.6, _CW * 0.22, _CW * 0.18],
        )
        y -= 8

    # ---- Disclaimer (short, single paragraph) ---------------------------
    disclaimer_text = (
        "Mutual fund investments are subject to market risks; read all scheme-related documents carefully. "
        "Past performance is not indicative of future results. Returns are computed using the XIRR methodology. "
        "This factsheet is for informational purposes only and does not constitute investment advice. "
        "Credent Asset Management Services Pvt Ltd  |  SEBI PMS Reg. INP000006101  |  www.credentglobal.com"
    )
    if y < 100:  # room check
        c.showPage()
        y = _PH - 30
    y = _draw_simple_section(c, y, "DISCLAIMER")
    _draw_paragraph(c, _LM, y - 4, _CW, disclaimer_text, font_size=7.5, leading=10)


def _draw_simple_section(c: PdfCanvas, y: float, title: str) -> float:
    """Small navy section-title bar."""
    h = 16
    c.setFillColor(_NAVY)
    c.rect(_LM, y - h, _CW, h, fill=True, stroke=False)
    c.setFillColor(_GOLD)
    c.rect(_LM, y - h, 3, h, fill=True, stroke=False)
    c.setFillColor(_WHITE)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(_LM + 10, y - 11, title)
    return y - h - 4


def _draw_two_col_table(c: PdfCanvas, y: float, rows: list[tuple[str, str]], label_w: float) -> float:
    row_h = 14
    for i, (label, value) in enumerate(rows):
        bg = _LIGHT_BLUE if i % 2 == 0 else _WHITE
        c.setFillColor(bg)
        c.rect(_LM, y - row_h, _CW, row_h, fill=True, stroke=False)
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.3)
        c.rect(_LM, y - row_h, _CW, row_h, fill=False, stroke=True)
        c.setFillColor(_TEXT_DARK)
        c.setFont("Helvetica-Bold", 7.5)
        c.drawString(_LM + 8, y - 10, label)
        c.setFont("Helvetica", 7.5)
        c.drawString(_LM + 8 + label_w, y - 10, str(value))
        y -= row_h
    return y


def _draw_three_col_table(
    c: PdfCanvas,
    y: float,
    headers: list[str],
    rows: list[tuple[str, str, str]],
    col_widths: list[float],
    total_last: bool = False,
) -> float:
    row_h = 15
    x_positions = [_LM]
    for w in col_widths[:-1]:
        x_positions.append(x_positions[-1] + w)
    # header
    c.setFillColor(_DARK_NAVY)
    c.rect(_LM, y - row_h, _CW, row_h, fill=True, stroke=False)
    c.setFillColor(_WHITE)
    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(x_positions[0] + 6, y - 10, headers[0])
    c.drawRightString(x_positions[1] + col_widths[1] - 6, y - 10, headers[1])
    c.drawRightString(x_positions[2] + col_widths[2] - 6, y - 10, headers[2])
    y -= row_h
    # rows
    for i, (a, b, d) in enumerate(rows):
        is_total = total_last and i == len(rows) - 1
        bg = _CREAM if is_total else (_LIGHT_BLUE if i % 2 == 0 else _WHITE)
        c.setFillColor(bg)
        c.rect(_LM, y - row_h, _CW, row_h, fill=True, stroke=False)
        c.setStrokeColor(_BORDER)
        c.setLineWidth(0.3)
        c.rect(_LM, y - row_h, _CW, row_h, fill=False, stroke=True)
        c.setFillColor(_TEXT_DARK)
        c.setFont("Helvetica-Bold" if is_total else "Helvetica", 7.5)
        c.drawString(x_positions[0] + 6, y - 10, str(a))
        c.drawRightString(x_positions[1] + col_widths[1] - 6, y - 10, str(b))
        c.drawRightString(x_positions[2] + col_widths[2] - 6, y - 10, str(d))
        y -= row_h
    return y


def _draw_paragraph(c: PdfCanvas, x: float, y: float, width: float, text: str, font_size: float = 7.5, leading: float = 10) -> float:
    """Simple, iterative word-wrap. No recursion, no beginText."""
    c.setFillColor(_TEXT_MED)
    c.setFont("Helvetica", font_size)
    words = text.split()
    line_words: list[str] = []
    for word in words:
        candidate = " ".join(line_words + [word])
        if c.stringWidth(candidate, "Helvetica", font_size) <= width - 12 or not line_words:
            line_words.append(word)
        else:
            c.drawString(x + 6, y, " ".join(line_words))
            y -= leading
            line_words = [word]
    if line_words:
        c.drawString(x + 6, y, " ".join(line_words))
        y -= leading
    return y



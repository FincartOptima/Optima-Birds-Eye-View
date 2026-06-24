from __future__ import annotations

from bisect import bisect_right
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

from create_client_factsheet_report import CATEGORY_ORDER

BASE_DIR = Path(__file__).resolve().parent

SOURCE_FILES = {
    'equity': BASE_DIR / 'Equity Source.xlsx',
    'debt':   BASE_DIR / 'Debt Source.xlsx',
    'gold':   BASE_DIR / 'Gold Source.xlsx',
    'cash':   BASE_DIR / 'Cash Source.xlsx',
}
SCHEME_CATEGORY_FILE = BASE_DIR / 'Scheme and Category.xlsx'

PERIOD_DAYS: list[tuple[str, int]] = [
    ('1M',   30),
    ('2M',   60),
    ('3M',   91),
    ('6M',  182),
    ('1Y',  365),
    ('2Y',  730),
    ('3Y', 1095),
    ('5Y', 1825),
]
PERIOD_LABELS = [p[0] for p in PERIOD_DAYS]


# ---------------------------------------------------------------------------
# NAV loading
# ---------------------------------------------------------------------------

def _load_active_fund_map() -> dict[str, str]:
    """Returns {fund_display_name: source_key} for Active='Y' funds only."""
    active: dict[str, str] = {}
    for source_key, fpath in SOURCE_FILES.items():
        if not fpath.exists():
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        if '_Codes' not in wb.sheetnames:
            wb.close()
            continue
        for row in wb['_Codes'].iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) > 2 and row[2] == 'Y':
                active[str(row[0]).strip()] = source_key
        wb.close()
    return active


def load_nav_index() -> dict[str, dict]:
    """
    Returns {fund_name: {'dates': [datetime, ...], 'values': [float, ...]}}
    Only Active='Y' funds are included.
    """
    active_map = _load_active_fund_map()
    nav_index: dict[str, dict] = {f: {'dates': [], 'values': []} for f in active_map}

    for source_key, fpath in SOURCE_FILES.items():
        if not fpath.exists():
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

        # columns that belong to active funds from this source
        fund_cols: dict[str, int] = {
            h: i for i, h in enumerate(headers)
            if h in active_map and active_map[h] == source_key
        }

        for row in ws.iter_rows(min_row=2, values_only=True):
            date = row[0]
            if not isinstance(date, datetime):
                continue
            for fund, col_idx in fund_cols.items():
                v = row[col_idx] if col_idx < len(row) else None
                if isinstance(v, (int, float)):
                    nav_index[fund]['dates'].append(date)
                    nav_index[fund]['values'].append(float(v))
        wb.close()

    return nav_index


def load_scheme_categories() -> dict[str, str]:
    """Returns {scheme_display_name: category}."""
    if not SCHEME_CATEGORY_FILE.exists():
        return {}
    wb = openpyxl.load_workbook(SCHEME_CATEGORY_FILE, data_only=True, read_only=True)
    ws = wb.active
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            mapping[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()
    return mapping


# ---------------------------------------------------------------------------
# Return computation
# ---------------------------------------------------------------------------

def _value_on_or_before(dates: list[datetime], values: list[float], target: datetime) -> float | None:
    pos = bisect_right(dates, target) - 1
    return values[pos] if pos >= 0 else None


def _compute_return(dates: list[datetime], values: list[float], report_date: datetime, days_back: int) -> float | None:
    from_date = report_date - timedelta(days=days_back)
    v_start = _value_on_or_before(dates, values, from_date)
    v_end   = _value_on_or_before(dates, values, report_date)
    if v_start and v_end and v_start > 0:
        return round((v_end / v_start - 1) * 100, 2)
    return None


def _bse_returns(bse_prices: list[tuple], report_date: datetime) -> dict[str, float | None]:
    dates  = [p[0] for p in bse_prices]
    values = [p[1] for p in bse_prices]
    return {label: _compute_return(dates, values, report_date, days) for label, days in PERIOD_DAYS}


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def get_master_data(reports: list, bse_prices: list[tuple], report_date: datetime) -> dict:
    nav_index   = load_nav_index()
    scheme_cats = load_scheme_categories()
    bse_ret     = _bse_returns(bse_prices, report_date)

    # ── Fund performance ────────────────────────────────────────────────────
    fund_perf: list[dict] = []
    for fund_name, idx in nav_index.items():
        returns = {
            label: _compute_return(idx['dates'], idx['values'], report_date, days)
            for label, days in PERIOD_DAYS
        }
        fund_perf.append({
            'name':     fund_name,
            'category': scheme_cats.get(fund_name, 'Other'),
            'returns':  returns,
        })
    fund_perf.sort(key=lambda x: (x['category'], x['name']))

    # ── Category performance (equal-weight average of fund returns) ─────────
    cat_buckets: dict[str, dict[str, list]] = defaultdict(lambda: {l: [] for l, _ in PERIOD_DAYS})
    for fp in fund_perf:
        for label, val in fp['returns'].items():
            if val is not None:
                cat_buckets[fp['category']][label].append(val)

    cat_perf: list[dict] = []
    for cat in sorted(cat_buckets):
        cat_perf.append({
            'name': cat,
            'returns': {
                label: round(sum(vals) / len(vals), 2) if vals else None
                for label, vals in cat_buckets[cat].items()
            },
        })

    # ── Cash in hand ────────────────────────────────────────────────────────
    active_reports = [r for r in reports if r.current_value > 0]
    total_cash  = sum(r.only_cash for r in active_reports)
    total_aum   = sum(r.current_value for r in active_reports)
    cash_pct    = round(total_cash / total_aum * 100, 2) if total_aum else 0.0

    per_client_cash = [
        {
            'client':          r.client_name,
            'cash_amount':     r.only_cash,
            'portfolio_value': r.current_value,
            'cash_pct':        round(r.only_cash / r.current_value * 100, 2) if r.current_value else 0.0,
        }
        for r in active_reports
    ]

    # ── Client × Fund matrix ────────────────────────────────────────────────
    # Collect all funds held by at least one active client
    fund_set: list[str] = []
    seen: set[str] = set()
    for r in active_reports:
        for h in sorted(r.holdings, key=lambda h: h.current_value, reverse=True):
            if h.current_value > 0 and h.scheme_name not in seen:
                fund_set.append(h.scheme_name)
                seen.add(h.scheme_name)

    client_names = [r.client_name for r in active_reports]

    fund_matrix_data: dict[str, list[float]] = {}
    for fund_name in fund_set:
        row: list[float] = []
        for r in active_reports:
            h = next((hh for hh in r.holdings if hh.scheme_name == fund_name), None)
            pct = round(h.current_value / r.current_value * 100, 2) if (h and r.current_value) else 0.0
            row.append(pct)
        fund_matrix_data[fund_name] = row

    # ── Client × Category matrix ─────────────────────────────────────────────
    cat_matrix_data: dict[str, list[float]] = {}
    for cat in CATEGORY_ORDER:
        row = []
        for r in active_reports:
            cat_row = next((cr for cr in r.category_rows if cr['Category'] == cat), None)
            pct = round(cat_row['Allocation %'] * 100, 2) if cat_row else 0.0
            row.append(pct)
        cat_matrix_data[cat] = row

    return {
        'fund_performance':   fund_perf,
        'category_performance': cat_perf,
        'bse_returns':        bse_ret,
        'period_labels':      PERIOD_LABELS,
        'cash_summary': {
            'total_cash':      total_cash,
            'total_aum':       total_aum,
            'cash_pct':        cash_pct,
            'per_client':      per_client_cash,
        },
        'client_fund_matrix': {
            'clients': client_names,
            'funds':   fund_set,
            'data':    fund_matrix_data,
        },
        'client_category_matrix': {
            'clients':    client_names,
            'categories': CATEGORY_ORDER,
            'data':       cat_matrix_data,
        },
        'report_date': report_date.strftime('%d %b %Y'),
    }
